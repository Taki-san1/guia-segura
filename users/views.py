# users/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView, PasswordResetView, PasswordChangeView
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.views import View
from django.contrib.auth.decorators import login_required
from django.db import models
import time
import os
import requests as req

# Scraping
from datetime import datetime
from bs4 import BeautifulSoup

# Modelos y Formularios
from django.views.generic import ListView
from django.contrib.auth.models import User
from django.contrib.auth.mixins import UserPassesTestMixin
from .forms import (
    FormularioRegistro,
    FormularioAcceso,
    FormularioActualizarUsuario,
    FormularioActualizarPerfil,
)
from .models import Perfil, HistorialGuia

from django.contrib.auth.decorators import user_passes_test

#excel
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.core.mail import EmailMessage
from io import BytesIO
from django.conf import settings


# ====================================================================
# 🔐 PERMISOS
# ====================================================================
def es_admin(user):
    return user.is_superuser or user.is_staff


# ====================================================================
# 🔥 PANEL DE USUARIOS
# ====================================================================
@user_passes_test(es_admin)
def panel_usuarios(request):
    usuarios = User.objects.filter(is_active=True).order_by("id")
    return render(request, "users/panel_usuarios.html", {"usuarios": usuarios})

@user_passes_test(es_admin)
def usuarios_inactivos(request):
    usuarios = User.objects.filter(is_active=False).order_by("id")
    return render(request, "users/usuarios_inactivos.html", {"usuarios": usuarios})

@user_passes_test(es_admin)
def activar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    usuario.is_active = True
    usuario.save()
    messages.success(request, f"El usuario '{usuario.username}' fue reactivado.")
    return redirect("usuarios_inactivos")


@user_passes_test(es_admin)
def crear_usuario(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        password = request.POST["password"]
        User.objects.create_user(username=username, email=email, password=password)
        messages.success(request, "Usuario creado correctamente")
        return redirect("panel_usuarios")
    return render(request, "users/crear_usuario.html")


@user_passes_test(es_admin)
def editar_usuario(request, user_id):
    usuario = User.objects.get(id=user_id)
    if request.method == "POST":
        usuario.username = request.POST["username"]
        usuario.email = request.POST["email"]
        if request.POST["password"]:
            usuario.set_password(request.POST["password"])
        usuario.save()
        messages.success(request, "Usuario actualizado")
        return redirect("panel_usuarios")
    return render(request, "users/editar_usuario.html", {"usuario": usuario})


@user_passes_test(es_admin)
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    if usuario == request.user:
        messages.error(request, "No puedes desactivar tu propia cuenta.")
        return redirect("panel_usuarios")
    usuario.is_active = False
    usuario.is_staff = False
    usuario.is_superuser = False
    usuario.save()
    messages.success(request, f"El usuario '{usuario.username}' fue desactivado correctamente.")
    return redirect("panel_usuarios")


# ====================================================================
# 🚀 CAMBIAR ROL DE USUARIO
# ====================================================================
@user_passes_test(es_admin)
def cambiar_rol_usuario(request, user_id, rol):
    usuario = get_object_or_404(User, id=user_id)
    if usuario == request.user:
        messages.error(request, "No puedes cambiar tu propio rol.")
        return redirect("panel_usuarios")
    if rol == "superuser":
        usuario.is_superuser = True
        usuario.is_staff = True
    elif rol == "staff":
        usuario.is_superuser = False
        usuario.is_staff = True
    elif rol == "normal":
        usuario.is_superuser = False
        usuario.is_staff = False
    else:
        messages.error(request, "Rol inválido.")
        return redirect("panel_usuarios")
    usuario.save()
    messages.success(request, f"El rol del usuario '{usuario.username}' fue actualizado a '{rol}'.")
    return redirect("panel_usuarios")


# ====================================================================
# HOME
# ====================================================================
def home(request):
    return render(request, 'users/home.html')


class VistaRegistro(View):
    form_class = FormularioRegistro
    initial = {'key': 'value'}
    template_name = 'users/register.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('/')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Cuenta creada")
            return redirect("login")
        return render(request, self.template_name, {'form': form})


@login_required
def profile(request):
    if request.method == 'POST':
        user_form = FormularioActualizarUsuario(request.POST, instance=request.user)
        profile_form = FormularioActualizarPerfil(request.POST, request.FILES, instance=request.user.perfil)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "Perfil actualizado")
            return redirect("users-profile")
    else:
        user_form = FormularioActualizarUsuario(instance=request.user)
        profile_form = FormularioActualizarPerfil(instance=request.user.perfil)

    contexto = {'user_form': user_form, 'profile_form': profile_form}

    if request.user.username == 'ADMIN1':
        return render(request, 'users/profile.html', contexto)
    else:
        return render(request, 'users/profile_usuario.html', contexto)


class VistaListaUsuarios(UserPassesTestMixin, ListView):
    model = User
    context_object_name = 'user_list'
    template_name = 'users/users_list.html'
    paginate_by = 15

    def test_func(self):
        return self.request.user.username == 'ADMIN1'


class VistaAccesoPersonalizada(LoginView):
    pass


class VistaRestablecerContrasena(SuccessMessageMixin, PasswordResetView):
    template_name = 'users/password_reset.html'
    email_template_name = 'users/password_reset_email.html'
    subject_template_name = 'users/password_reset_subject.txt'
    success_message = "Se enviaron las instrucciones por correo"


class VistaCambiarContrasena(SuccessMessageMixin, PasswordChangeView):
    template_name = 'users/change_password.html'
    success_url = reverse_lazy('users-profile')
    success_message = "Contraseña cambiada"


# ====================================================================
# 🔍 CONSULTAR GUÍA — via servidor local con ngrok
# ====================================================================
@login_required
def VistaConsultarGuia(request):
    resultados_consulta = None
    guia_consultada = None

    if request.method == 'POST':
        guia_consultada = request.POST.get('guia_a_consultar', '').strip()

        if guia_consultada:
            ultimo = HistorialGuia.objects.aggregate(models.Max('consulta_id')).get("consulta_id__max")
            nuevo_consulta_id = (ultimo or 0) + 1

            try:
                scraper_url = os.getenv('SCRAPER_URL')
                scraper_secret = os.getenv('SCRAPER_SECRET')

                if not scraper_url:
                    messages.error(request, "El servidor de consultas no está configurado.")
                    return render(request, "users/consultar_guia.html", {
                        "resultados": None,
                        "guia_consultada": guia_consultada,
                    })

                response = req.get(
                    f"{scraper_url}/scrape",
                    params={"guia": guia_consultada},
                    headers={"X-API-Secret": scraper_secret},
                    timeout=60,
                )

                if response.status_code == 200:
                    data = response.json()

                    if data.get("success") and data.get("eventos"):
                        resultados_consulta = data["eventos"]

                        for evento in resultados_consulta:
                            HistorialGuia.objects.create(
                                usuario=request.user,
                                consulta_id=nuevo_consulta_id,
                                numero_guia=guia_consultada,
                                fecha=evento["fecha"],
                                hora=evento["hora"],
                                estado=evento["estado"],
                                sucursal=evento["sucursal"],
                                fecha_consulta=datetime.now()
                            )
                        messages.success(request, "Guía consultada correctamente")
                    else:
                        messages.warning(request, data.get("error", "No se encontraron eventos para esta guía."))
                else:
                    messages.error(request, f"Error al consultar la guía. Código: {response.status_code}")

            except Exception as e:
                messages.error(request, f"Error: {e}")

    return render(request, "users/consultar_guia.html", {
        "resultados": resultados_consulta,
        "guia_consultada": guia_consultada,
    })


# ====================================================================
# 📦 PANEL DE GUÍAS
# ====================================================================
@user_passes_test(es_admin)
def panel_guias(request):
    consultas = (
        HistorialGuia.objects
        .filter(activo=True)
        .values("consulta_id", "numero_guia", "usuario_id")
        .annotate(total_eventos=models.Count("id"))
        .order_by("-consulta_id")
    )
    return render(request, "users/guias_panel.html", {"consultas": consultas})


# ====================================================================
# 📄 DETALLE DE CONSULTA
# ====================================================================
@user_passes_test(es_admin)
def detalle_consulta(request, consulta_id):
    eventos = (
        HistorialGuia.objects
        .filter(consulta_id=consulta_id, activo=True)
        .order_by('id')
    )
    if not eventos.exists():
        messages.error(request, "No hay eventos activos para esta consulta.")
        return redirect("panel_guias")
    return render(request, "users/guias_detalle.html", {
        "consulta_id": consulta_id,
        "numero_guia": eventos.first().numero_guia,
        "eventos": eventos
    })


# ====================================================================
# ➕ CREAR EVENTO
# ====================================================================
@user_passes_test(es_admin)
def crear_evento(request, consulta_id):
    consulta = HistorialGuia.objects.filter(consulta_id=consulta_id).first()
    if not consulta:
        messages.error(request, "Consulta no encontrada.")
        return redirect("panel_guias")

    if request.method == "POST":
        estado = request.POST.get("estado")
        if estado == "otro":
            estado = request.POST.get("estado_otro")

        HistorialGuia.objects.create(
            usuario=request.user,
            consulta_id=consulta_id,
            numero_guia=consulta.numero_guia,
            fecha=request.POST.get("fecha"),
            hora=request.POST.get("hora"),
            estado=estado,
            sucursal=request.POST.get("sucursal"),
            fecha_consulta=datetime.now()
        )
        messages.success(request, "Evento creado correctamente.")
        return redirect("detalle_consulta", consulta_id=consulta_id)

    return render(request, "users/guias_crear_evento.html", {
        "consulta_id": consulta_id,
        "numero_guia": consulta.numero_guia,
    })


# ====================================================================
# ✏ EDITAR EVENTO
# ====================================================================
@user_passes_test(es_admin)
def editar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)

    if request.method == "POST":
        estado = request.POST.get("estado")
        if estado == "otro":
            estado = request.POST.get("estado_otro")

        evento.fecha = request.POST.get("fecha")
        evento.hora = request.POST.get("hora")
        evento.estado = estado
        evento.sucursal = request.POST.get("sucursal")
        evento.save()

        messages.success(request, "Evento actualizado correctamente.")
        return redirect("detalle_consulta", consulta_id=consulta_id)

    return render(request, "users/guias_editar_evento.html", {
        "evento": evento,
        "consulta_id": consulta_id,
        "numero_guia": evento.numero_guia,
    })


# ====================================================================
# 🗑 ELIMINAR EVENTO (borrado lógico)
# ====================================================================
@user_passes_test(es_admin)
def eliminar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)
    evento.activo = False
    evento.save()
    messages.success(request, "Evento marcado como eliminado.")
    return redirect("detalle_consulta", consulta_id=consulta_id)


# ====================================================================
# 🔄 INACTIVOS Y RESTAURAR
# ====================================================================
@user_passes_test(es_admin)
def detalle_consulta_inactivos(request, consulta_id):
    eventos = HistorialGuia.objects.filter(consulta_id=consulta_id, activo=False).order_by("id")
    return render(request, "users/guias_detalle_inactivos.html", {
        "eventos": eventos,
        "consulta_id": consulta_id
    })


@user_passes_test(es_admin)
def restaurar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)
    evento.activo = True
    evento.save()
    messages.success(request, "Evento restaurado correctamente.")
    return redirect("detalle_consulta_inactivos", consulta_id=consulta_id)


# ====================================================================
# 📊 EXPORTAR EXCEL
# ====================================================================
@user_passes_test(es_admin)
def exportar_guias_excel(request):
    registros = HistorialGuia.objects.filter(activo=True).order_by(
        "numero_guia", "consulta_id", "fecha_consulta"
    )

    if not registros.exists():
        messages.warning(request, "No hay registros activos para exportar.")
        return redirect("panel_guias")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guías"

    columnas = [
        "ID Consulta", "Número de guía", "Usuario",
        "Fecha evento", "Hora", "Estado",
        "Sucursal / ciudad", "Fecha de consulta",
    ]

    for col_num, titulo in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=titulo)

    fila = 2
    for r in registros:
        ws.cell(row=fila, column=1, value=r.consulta_id)
        ws.cell(row=fila, column=2, value=r.numero_guia)
        ws.cell(row=fila, column=3, value=r.usuario.username if r.usuario else "")
        ws.cell(row=fila, column=4, value=r.fecha)
        ws.cell(row=fila, column=5, value=r.hora)
        ws.cell(row=fila, column=6, value=r.estado)
        ws.cell(row=fila, column=7, value=r.sucursal)
        ws.cell(row=fila, column=8, value=r.fecha_consulta.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_consulta else "")
        fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()

    if request.user.email:
        try:
            mensaje = EmailMessage(
                subject="Reporte de guías generado",
                body=f"Hola {request.user.username},\n\nAdjunto encontrarás el reporte de guías.",
                from_email=settings.EMAIL_HOST_USER,
                to=[request.user.email],
            )
            mensaje.attach(
                "reporte_guias.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            mensaje.send()
            messages.success(request, f"Reporte enviado a {request.user.email}")
        except Exception as e:
            messages.error(request, f"Error enviando correo: {e}")
    else:
        messages.warning(request, "El usuario no tiene correo configurado.")

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_guias.xlsx"'
    return response